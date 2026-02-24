# -*- coding: utf-8 -*-
"""
===========================================================
  Summit Logic V3 - 스마트스토어 × 대한통운 LOIS 자동화 도구
===========================================================
[실행 방법]
  pip install streamlit pandas openpyxl msoffcrypto-tool
  streamlit run test_run.py

[V3 기능]
  탭 1: 스마트스토어 주문서 → 대한통운 LOIS 접수 파일 변환
        (합배송 자동 감지 + 데이터 정제)
  탭 2: 대한통운 LOIS 결과 + 스마트스토어 원본 → 송장번호 자동 기입
        (합배송 묶음 전체에 동일 송장번호 입력)
  보안: Access Key 입력 시에만 기능 활성화
===========================================================
"""

import io
import re
import msoffcrypto
import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# ===========================================================
# 상수: 네이버 스마트스토어 엑셀 컬럼 인덱스 (0-based)
# ===========================================================
NAVER = {
    "상품주문번호":  0,   # A열
    "택배사":       7,   # H열
    "송장번호":     8,   # I열
    "수취인명":     13,  # N열
    "상품명":       20,  # U열
    "수량":         26,  # AA열
    "수취인연락처1": 48,  # AW열
    "합배송지":     50,  # AY열
    "우편번호":     54,  # BC열
    "배송메세지":   55,  # BD열
}

# 사이트 접근 제어 키
ACCESS_KEY = "summit2026"

# CJ LOIS 주소 필드 최대 길이
ADDRESS_MAX_LEN = 100


# ===========================================================
# 유틸 함수
# ===========================================================

# ===========================================================
# [V3 신규] 데이터 정제 함수
# ===========================================================

_EMOJI_RE = re.compile(
    "["
    "\U0001F000-\U0001FFFF"
    "\U00002600-\U000027BF"
    "\U0000200B-\U0000200F"
    "\U0000FE00-\U0000FE0F"
    "]+",
    flags=re.UNICODE,
)


def clean_text(text: str) -> str:
    """이름·주소·배송메시지에서 이모지·제어문자를 제거합니다."""
    text = _EMOJI_RE.sub("", str(text))
    text = re.sub(r"[\x00-\x1f\x7f]", " ", text)
    text = re.sub(r" {2,}", " ", text)
    return text.strip()


def clean_phone(phone: str) -> str:
    """전화번호에서 숫자 이외의 모든 문자(하이픈 등)를 제거합니다."""
    return re.sub(r"[^0-9]", "", str(phone))


def truncate_address(address: str, max_len: int = ADDRESS_MAX_LEN) -> str:
    """주소가 CJ LOIS 업로드 길이 제한을 초과하면 잘라냅니다."""
    return address[:max_len] if len(address) > max_len else address


# ===========================================================
# 유틸 함수
# ===========================================================

def find_header_row(file_obj) -> int:
    """'상품주문번호' 텍스트와 정확히 일치하는 셀이 있는 행 번호를 반환합니다."""
    file_obj.seek(0)
    df_raw = pd.read_excel(file_obj, header=None, dtype=str)
    for idx, row in df_raw.iterrows():
        if (row.astype(str).str.strip() == "상품주문번호").any():
            return int(idx)
    raise ValueError(
        "'상품주문번호' 컬럼을 찾을 수 없습니다.\n"
        "네이버 스마트스토어에서 다운로드한 원본 엑셀 파일인지 확인해 주세요."
    )


def read_naver_excel(file_obj) -> pd.DataFrame:
    """네이버 스마트스토어 주문 엑셀을 읽고 불량 행을 제거합니다."""
    header_row = find_header_row(file_obj)
    file_obj.seek(0)
    df = pd.read_excel(file_obj, header=header_row, dtype=str)
    df = df.fillna("")
    order_col = df.columns[0]
    df = df[
        (df[order_col].str.strip() != "") &
        (df[order_col].str.strip() != "상품주문번호")
    ].reset_index(drop=True)
    return df


def build_cj_upload_df(df_smart: pd.DataFrame) -> tuple:
    """
    [V3] 스마트스토어 → CJ LOIS 접수 양식 변환 (정제 + 합배송)
    반환: (변환 DataFrame, 원본 주문 건수)
    """
    df = pd.DataFrame({
        "고객주문번호": df_smart.iloc[:, NAVER["상품주문번호"]].str.strip(),
        "수취인명":     df_smart.iloc[:, NAVER["수취인명"]].apply(lambda x: clean_text(str(x))),
        "연락처":       df_smart.iloc[:, NAVER["수취인연락처1"]].apply(lambda x: clean_phone(str(x))),
        "우편번호":     df_smart.iloc[:, NAVER["우편번호"]].str.strip(),
        "주소":         df_smart.iloc[:, NAVER["합배송지"]].apply(
                            lambda x: truncate_address(clean_text(str(x)))),
        "상품명":       df_smart.iloc[:, NAVER["상품명"]].str.strip(),
        "수량":         df_smart.iloc[:, NAVER["수량"]].str.strip(),
        "배송메시지":   df_smart.iloc[:, NAVER["배송메세지"]].apply(lambda x: clean_text(str(x))),
    })
    df = df[df["고객주문번호"] != ""].reset_index(drop=True)
    original_count = len(df)

    rows = []
    for (name, phone, addr), group in df.groupby(
        ["수취인명", "연락처", "주소"], sort=False
    ):
        first = group.iloc[0]
        products = group["상품명"].tolist()
        product_summary = (
            products[0]
            if len(products) == 1
            else f"{products[0]} 외 {len(products) - 1}건"
        )
        try:
            qty_list = [int(q) for q in group["수량"] if str(q).strip().isdigit()]
            total_qty = sum(qty_list) if qty_list else first["수량"]
        except Exception:
            total_qty = first["수량"]

        rows.append({
            "고객주문번호": first["고객주문번호"],
            "수취인명":     name,
            "연락처":       phone,
            "우편번호":     first["우편번호"],
            "주소":         addr,
            "상품명":       product_summary,
            "수량":         str(total_qty),
            "배송메시지":   first["배송메시지"],
        })

    return pd.DataFrame(rows), original_count


def match_and_fill_waybill(smart_file_obj, cj_df: pd.DataFrame):
    """
    [V3 템플릿 유지형 + 합배송 대응 송장 매칭]
    그룹 대표 주문번호로 CJ 조회 후, 묶음 내 모든 행에 동일 송장번호 기입.
    반환: (엑셀 바이트, 매칭 성공 건수, 미발급 건수, 미발급 주문번호 목록)
    """
    cj_lookup: dict = {}
    for _, row in cj_df.iterrows():
        key = str(row.get("고객주문번호", "")).strip()
        val = str(row.get("운송장번호", "")).strip()
        if key and key not in cj_lookup:
            cj_lookup[key] = val

    df_smart = read_naver_excel(smart_file_obj)
    clean_keys = pd.DataFrame({
        "order_no": df_smart.iloc[:, NAVER["상품주문번호"]].str.strip(),
        "name":     df_smart.iloc[:, NAVER["수취인명"]].apply(lambda x: clean_text(str(x))),
        "phone":    df_smart.iloc[:, NAVER["수취인연락처1"]].apply(lambda x: clean_phone(str(x))),
        "addr":     df_smart.iloc[:, NAVER["합배송지"]].apply(
                        lambda x: truncate_address(clean_text(str(x)))),
    })
    clean_keys = clean_keys[clean_keys["order_no"] != ""].reset_index(drop=True)

    rep_of: dict = {}
    for _, group in clean_keys.groupby(["name", "phone", "addr"], sort=False):
        orders = group["order_no"].tolist()
        rep = orders[0]
        for o in orders:
            rep_of[o] = rep

    order_to_waybill: dict = {}
    for order_no, rep in rep_of.items():
        waybill = cj_lookup.get(rep, "") or cj_lookup.get(order_no, "")
        if waybill:
            order_to_waybill[order_no] = waybill

    header_idx = find_header_row(smart_file_obj)
    data_start_row = header_idx + 2

    smart_file_obj.seek(0)
    wb = load_workbook(smart_file_obj)
    ws = wb.active

    matched = 0
    unmatched = 0
    unmatched_list: list = []

    for row_cells in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
        order_no = str(row_cells[NAVER["상품주문번호"]].value or "").strip()
        if not order_no:
            continue
        waybill = order_to_waybill.get(order_no, "")
        if waybill:
            row_cells[NAVER["택배사"]].value   = "CJ대한통운"
            row_cells[NAVER["송장번호"]].value = waybill
            matched += 1
        else:
            row_cells[NAVER["택배사"]].value   = "미발급"
            row_cells[NAVER["송장번호"]].value = "미발급"
            unmatched += 1
            unmatched_list.append(order_no)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), matched, unmatched, unmatched_list


def unlock_excel(file_obj, password: str = "") -> io.BytesIO:
    """엑셀 암호 해제. 비밀번호 없으면 그대로 BytesIO 반환."""
    file_obj.seek(0)
    raw = file_obj.read()
    if not password.strip():
        return io.BytesIO(raw)
    encrypted_buf = io.BytesIO(raw)
    office_file = msoffcrypto.OfficeFile(encrypted_buf)
    office_file.load_key(password=password.strip())
    decrypted_buf = io.BytesIO()
    office_file.decrypt(decrypted_buf)
    decrypted_buf.seek(0)
    return decrypted_buf


def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    """데이터프레임을 엑셀 바이트 스트림으로 변환."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    buf.seek(0)
    return buf.getvalue()


# ===========================================================
# Streamlit UI
# ===========================================================

st.set_page_config(
    page_title="Summit Logic",
    page_icon="📦",
    layout="centered",
)

# ── [V3] 사이드바: Access Key ──
with st.sidebar:
    st.markdown("### 🔐 Access Control")
    st.markdown("---")
    access_input = st.text_input(
        "Access Key",
        type="password",
        placeholder="접속 키를 입력하세요",
        key="access_key",
    )
    if access_input == ACCESS_KEY:
        st.success("✅ 인증 완료")
    elif access_input:
        st.error("❌ 잘못된 접속 키")
    else:
        st.info("키를 입력하면 기능이 활성화됩니다")
    st.markdown("---")
    st.caption("Summit Logic V3")

# ── 앱 헤더 ──
st.title("📦 Summit Logic")
st.caption("스마트스토어 × 대한통운 LOIS 자동화 도구 V3")
st.divider()

# ── [V3] Access Key 게이트 ──
if access_input != ACCESS_KEY:
    st.warning("🔐 좌측 사이드바에 **Access Key**를 입력해야 기능을 사용할 수 있습니다.")
    st.stop()

# ── 두 개의 탭 생성 ──
tab1, tab2 = st.tabs(["  📋 1. 접수 파일 생성  ", "  🔗 2. 송장 번호 매칭  "])


# ===========================================================
# 탭 1: 접수 파일 생성
# ===========================================================
with tab1:

    st.subheader("대한통운 LOIS 접수 파일 생성")
    st.info(
        "**네이버 스마트스토어 주문서**를 올리면 "
        "CJ 대한통운 LOIS 업로드 전용 양식으로 변환해 줍니다.\n\n"
        "스마트스토어 > 발주(주문)확인/발송관리 > 엑셀 다운로드 파일을 사용하세요."
    )

    uploaded_smart_t1 = st.file_uploader(
        "스마트스토어 주문서 (xlsx)",
        type=["xlsx"],
        key="tab1_uploader",
    )
    pw_t1 = st.text_input(
        "Excel Password (Optional)",
        type="password",
        key="tab1_pw",
        placeholder="엑셀 파일에 비밀번호가 있는 경우에만 입력하세요",
    )

    if uploaded_smart_t1:
        try:
            unlocked_t1 = unlock_excel(uploaded_smart_t1, pw_t1)
            df_smart = read_naver_excel(unlocked_t1)

            # [V3] build_cj_upload_df는 (df, original_count) 튜플 반환
            df_cj_upload, original_count = build_cj_upload_df(df_smart)
            total = len(df_cj_upload)
            bundled = original_count - total

            m1, m2, m3 = st.columns(3)
            m1.metric("원본 주문 건수", f"{original_count}건")
            m2.metric("발송 건수", f"{total}건")
            if bundled > 0:
                m3.metric("합배송 절약", f"{bundled}건")
                st.info(
                    f"🔗 **합배송 {bundled}건 자동 감지** — 수취인·연락처·주소가 동일한 주문을 "
                    "1건으로 묶었습니다. 상품명은 `상품A 외 N건` 형태로 요약되었습니다."
                )

            with st.expander("컬럼 매핑 확인"):
                st.table(pd.DataFrame({
                    "스마트스토어 컬럼명": [
                        "상품주문번호(A열)", "수취인명(N열)", "수취인연락처1(AW열)",
                        "우편번호(BC열)", "합배송지(AY열)", "상품명(U열)",
                        "수량(AA열)", "배송메세지(BD열)",
                    ],
                    "→ CJ LOIS 컬럼명": [
                        "고객주문번호", "수취인명(이모지제거)", "연락처(숫자만)",
                        "우편번호", "주소(100자제한)", "상품명(합배송요약)",
                        "수량(합산)", "배송메시지(이모지제거)",
                    ],
                }))

            st.dataframe(df_cj_upload, use_container_width=True)

            st.download_button(
                label="⬇ CJ LOIS 접수 파일 다운로드 (xlsx)",
                data=df_to_excel_bytes(df_cj_upload, "LOIS_접수"),
                file_name="CJ_LOIS_접수.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        except IndexError:
            st.error("파일 컬럼 구조가 예상과 다릅니다. 네이버 스마트스토어 원본 엑셀 파일인지 확인해 주세요.")
        except Exception as e:
            st.error(f"파일 처리 중 오류가 발생했습니다: {e}")
            with st.expander("오류 상세 내용"):
                st.exception(e)


# ===========================================================
# 탭 2: 송장 번호 매칭
#   스마트스토어 원본 + 대한통운 LOIS 결과 → 송장번호 자동 기입
#   [템플릿 유지형]: 1·2행 안내 문구를 그대로 보존하고 데이터만 수정
# ===========================================================
with tab2:

    st.subheader("대한통운 → 스마트스토어 송장번호 자동 매칭")
    st.info(
        "두 파일을 올리면 **상품주문번호 ↔ 고객주문번호**를 기준으로 "
        "자동으로 매칭해서 H열(택배사)과 I열(송장번호)을 채워 줍니다.\n\n"
        "원본 파일의 1·2행 양식(안내 문구)이 **그대로 유지**됩니다."
    )

    # ── 두 파일 업로드 (좌우 배치) ──
    col_left, col_right = st.columns(2)
    with col_left:
        uploaded_smart_t2 = st.file_uploader(
            "① 스마트스토어 원본 파일 (xlsx)",
            type=["xlsx"],
            key="tab2_smart",
            help="네이버 스마트스토어 주문 엑셀 원본 파일",
        )
        pw_t2 = st.text_input(
            "Excel Password (Optional)",
            type="password",
            key="tab2_pw",
            placeholder="비밀번호가 있는 경우에만 입력",
        )
    with col_right:
        uploaded_cj_t2 = st.file_uploader(
            "② 대한통운 LOIS 결과 파일 (xlsx)",
            type=["xlsx"],
            key="tab2_cj",
            help="대한통운 LOIS에서 운송장 발급 후 다운로드한 결과 파일",
        )

    # 두 파일이 모두 업로드됐을 때만 처리
    if uploaded_smart_t2 and uploaded_cj_t2:
        try:
            # ── CJ 파일 읽기 ──
            # 고객주문번호, 운송장번호 컬럼이 있는지 검증
            df_cj = pd.read_excel(uploaded_cj_t2, dtype=str).fillna("")

            required_cj_cols = ["고객주문번호", "운송장번호"]
            missing_cols = [c for c in required_cj_cols if c not in df_cj.columns]
            if missing_cols:
                st.error(
                    f"대한통운 파일에 필수 컬럼이 없습니다: **{missing_cols}**\n\n"
                    f"실제 컬럼 목록: `{list(df_cj.columns)}`"
                )
                st.stop()

            # ── 암호 해제 후 매칭 실행 (템플릿 유지형) ──
            with st.spinner("매칭 처리 중..."):
                unlocked_smart_t2 = unlock_excel(uploaded_smart_t2, pw_t2)
                result_bytes, matched, unmatched, unmatched_list = match_and_fill_waybill(
                    smart_file_obj=unlocked_smart_t2,
                    cj_df=df_cj,
                )

            total = matched + unmatched

            # ── 매칭 결과 통계 ──
            st.markdown("---")
            m1, m2, m3 = st.columns(3)
            m1.metric("전체 주문", f"{total}건")
            m2.metric("매칭 성공", f"{matched}건")
            m3.metric("미발급", f"{unmatched}건")

            if unmatched > 0:
                st.warning(
                    f"아래 **{unmatched}건**은 대한통운 파일에서 운송장번호를 찾지 못해 "
                    "'미발급'으로 표시되었습니다."
                )
                st.code("\n".join(unmatched_list), language=None)
            else:
                st.success("모든 주문의 송장번호가 성공적으로 매칭되었습니다!")

            # ── 매칭 결과 미리보기 (합배송 대응) ──
            st.markdown("**매칭 결과 미리보기**")
            header_row_prev = find_header_row(unlocked_smart_t2)
            unlocked_smart_t2.seek(0)
            df_preview = pd.read_excel(
                unlocked_smart_t2, header=header_row_prev, dtype=str
            ).fillna("")

            cj_preview_lookup = dict(
                zip(df_cj["고객주문번호"].str.strip(), df_cj["운송장번호"].str.strip())
            )

            # 합배송 그룹 대표 주문번호 맵 (미리보기에도 적용)
            ck = pd.DataFrame({
                "order_no": df_preview.iloc[:, NAVER["상품주문번호"]].str.strip(),
                "name":     df_preview.iloc[:, NAVER["수취인명"]].apply(lambda x: clean_text(str(x))),
                "phone":    df_preview.iloc[:, NAVER["수취인연락처1"]].apply(lambda x: clean_phone(str(x))),
                "addr":     df_preview.iloc[:, NAVER["합배송지"]].apply(
                                lambda x: truncate_address(clean_text(str(x)))),
            })
            rep_prev: dict = {}
            for _, grp in ck.groupby(["name", "phone", "addr"], sort=False):
                ords = grp["order_no"].tolist(); r = ords[0]
                for o in ords: rep_prev[o] = r

            preview_df = df_preview.iloc[:, [
                NAVER["상품주문번호"], NAVER["수취인명"],
                NAVER["상품명"], NAVER["택배사"], NAVER["송장번호"],
            ]].copy()
            preview_df.columns = ["상품주문번호", "수취인명", "상품명", "택배사", "송장번호"]
            preview_df = preview_df[preview_df["상품주문번호"].str.strip() != ""].copy()

            for idx, row in preview_df.iterrows():
                key = str(row["상품주문번호"]).strip()
                rep = rep_prev.get(key, key)
                wb_no = cj_preview_lookup.get(rep, "") or cj_preview_lookup.get(key, "")
                preview_df.at[idx, "택배사"]  = "CJ대한통운" if wb_no else "미발급"
                preview_df.at[idx, "송장번호"] = wb_no if wb_no else "미발급"

            st.dataframe(preview_df, use_container_width=True)

            # ── 다운로드 버튼 ──
            st.markdown("---")
            st.download_button(
                label="⬇ 송장번호 기입 완료된 스마트스토어 파일 다운로드 (xlsx)",
                data=result_bytes,
                file_name="스마트스토어_송장입력완료.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.caption(
                "다운로드한 파일을 스마트스토어 "
                "'발주(주문)확인/발송관리 > 일괄발송 처리' 메뉴에서 업로드하세요."
            )

        except Exception as e:
            st.error(f"파일 처리 중 오류가 발생했습니다: {e}")
            with st.expander("오류 상세 내용"):
                st.exception(e)

    elif uploaded_smart_t2 and not uploaded_cj_t2:
        st.info("대한통운 LOIS 결과 파일(②)도 업로드해 주세요.")
    elif not uploaded_smart_t2 and uploaded_cj_t2:
        st.info("스마트스토어 원본 파일(①)도 업로드해 주세요.")
