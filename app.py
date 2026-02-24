# -*- coding: utf-8 -*-
"""
===========================================================
  Summit Logic V3 - 스마트스토어 × 대한통운 LOIS 자동화 도구
===========================================================
[실행]  streamlit run app.py
[배포]  summitlogic.streamlit.app
[V3]   합배송 지능형 처리 / 데이터 정제 / 사이트 보안 잠금
===========================================================
"""

import io
import re
import msoffcrypto
import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# ===========================================================
# 상수: 네이버 스마트스토어 엑셀 컬럼 인덱스 (0-based)
# ===========================================================
NAVER = {
    "상품주문번호":  0,   # A열
    "택배사":       7,   # H열
    "송장번호":     8,   # I열
    "수취인명":     13,  # N열
    "상품명":       20,  # U열
    "수량":         26,  # AA열
    "수취인연락처1": 48,  # AW열
    "합배송지":     50,  # AY열
    "우편번호":     54,  # BC열
    "배송메세지":   55,  # BD열
}

# ── 보안: 사이트 접근 제어 키 (운영 환경에서는 st.secrets 로 교체 권장) ──
ACCESS_KEY = "summit2026"

# ── CJ LOIS 주소 필드 최대 길이 ──
ADDRESS_MAX_LEN = 100


# ===========================================================
# [V3 신규] 데이터 정제 함수
# ===========================================================

# 이모지 및 기타 기호 범위 (유니코드 블록 기준)
_EMOJI_RE = re.compile(
    "["
    "\U0001F000-\U0001FFFF"   # Misc Symbols, Emoticons, Transport, etc.
    "\U00002600-\U000027BF"   # Misc Symbols, Dingbats
    "\U0000200B-\U0000200F"   # Zero-width chars (ZWSP, ZWNJ, ZWJ, LRM, RLM)
    "\U0000FE00-\U0000FE0F"   # Variation Selectors
    "]+",
    flags=re.UNICODE,
)


def clean_text(text: str) -> str:
    """
    이름·주소·배송메시지에서 이모지 및 제어문자를 제거합니다.
    한글, 영문, 숫자, 공백, 기본 구두점(-.,()/)은 그대로 보존합니다.
    """
    text = _EMOJI_RE.sub("", str(text))
    # 탭·줄바꿈 등 제어문자를 공백으로 치환
    text = re.sub(r"[\x00-\x1f\x7f]", " ", text)
    # 연속된 공백을 하나로 압축
    text = re.sub(r" {2,}", " ", text)
    return text.strip()


def clean_phone(phone: str) -> str:
    """
    전화번호에서 하이픈·공백·괄호 등 숫자 이외의 모든 문자를 제거합니다.
    예) 010-1234-5678  →  01012345678
    """
    return re.sub(r"[^0-9]", "", str(phone))


def truncate_address(address: str, max_len: int = ADDRESS_MAX_LEN) -> str:
    """
    주소가 CJ LOIS 업로드 길이 제한(기본 100자)을 초과하면 잘라냅니다.
    """
    return address[:max_len] if len(address) > max_len else address


# ===========================================================
# 유틸 함수 (V2.2 헤더 탐색 로직 유지)
# ===========================================================

def find_header_row(file_obj) -> int:
    """
    '상품주문번호' 텍스트와 정확히 일치하는 셀이 있는 행 번호(0-based)를 반환합니다.

    contains() 대신 == 비교를 사용해, Row 0 안내 문구에 '상품주문번호'가
    설명 텍스트로 포함된 경우에도 헤더를 잘못 잡지 않습니다.
    """
    file_obj.seek(0)
    df_raw = pd.read_excel(file_obj, header=None, dtype=str)
    for idx, row in df_raw.iterrows():
        if (row.astype(str).str.strip() == "상품주문번호").any():
            return int(idx)
    raise ValueError(
        "'상품주문번호' 컬럼을 찾을 수 없습니다.\n"
        "네이버 스마트스토어에서 다운로드한 원본 엑셀 파일인지 확인해 주세요."
    )


def read_naver_excel(file_obj) -> pd.DataFrame:
    """
    네이버 스마트스토어 주문 엑셀을 안전하게 읽습니다.

    1. '상품주문번호'가 정확히 일치하는 행을 헤더로 동적 탐색
    2. 빈 행 / 중복 헤더 잔재 행 제거
    3. dtype=str → 주문번호·전화번호 앞자리 0 보존
    """
    header_row = find_header_row(file_obj)
    file_obj.seek(0)
    df = pd.read_excel(file_obj, header=header_row, dtype=str)
    df = df.fillna("")
    order_col = df.columns[0]
    df = df[
        (df[order_col].str.strip() != "") &
        (df[order_col].str.strip() != "상품주문번호")
    ].reset_index(drop=True)
    return df


def build_cj_upload_df(df_smart: pd.DataFrame) -> tuple:
    """
    [V3] 스마트스토어 데이터프레임 → CJ 대한통운 LOIS 접수 양식 변환

    변경사항 (V3):
    1. 데이터 정제
       - 이름·주소·배송메시지: 이모지·제어문자 제거 (clean_text)
       - 전화번호: 숫자만 추출 (clean_phone)
       - 주소: 길이 초과 시 잘라냄 (truncate_address)
    2. 합배송(Bundling) 처리
       - 수취인명 + 연락처 + 주소가 동일한 주문을 1건으로 묶음
       - 상품명: "상품A 외 N건" 형태로 요약
       - 수량: 그룹 내 합산
       - 고객주문번호: 그룹의 첫 번째 상품주문번호 (Tab 2 매칭 키)

    반환: (변환 DataFrame, 원본 주문 건수)
    """
    # ── 1) 추출 + 정제 ──
    df = pd.DataFrame({
        "고객주문번호": df_smart.iloc[:, NAVER["상품주문번호"]].str.strip(),
        "수취인명":     df_smart.iloc[:, NAVER["수취인명"]].apply(
                            lambda x: clean_text(str(x))),
        "연락처":       df_smart.iloc[:, NAVER["수취인연락처1"]].apply(
                            lambda x: clean_phone(str(x))),
        "우편번호":     df_smart.iloc[:, NAVER["우편번호"]].str.strip(),
        "주소":         df_smart.iloc[:, NAVER["합배송지"]].apply(
                            lambda x: truncate_address(clean_text(str(x)))),
        "상품명":       df_smart.iloc[:, NAVER["상품명"]].str.strip(),
        "수량":         df_smart.iloc[:, NAVER["수량"]].str.strip(),
        "배송메시지":   df_smart.iloc[:, NAVER["배송메세지"]].apply(
                            lambda x: clean_text(str(x))),
    })
    df = df[df["고객주문번호"] != ""].reset_index(drop=True)
    original_count = len(df)

    # ── 2) 합배송 그룹핑: 수취인명 + 연락처 + 주소 기준 ──
    rows = []
    for (name, phone, addr), group in df.groupby(
        ["수취인명", "연락처", "주소"], sort=False
    ):
        first = group.iloc[0]
        products = group["상품명"].tolist()

        # 상품명 요약: 1건이면 그대로, 2건 이상이면 "상품A 외 N건"
        product_summary = (
            products[0]
            if len(products) == 1
            else f"{products[0]} 외 {len(products) - 1}건"
        )

        # 수량 합산 (숫자 변환 불가 시 첫 번째 값 사용)
        try:
            qty_list = [int(q) for q in group["수량"] if str(q).strip().isdigit()]
            total_qty = sum(qty_list) if qty_list else first["수량"]
        except Exception:
            total_qty = first["수량"]

        rows.append({
            "고객주문번호": first["고객주문번호"],
            "수취인명":     name,
            "연락처":       phone,
            "우편번호":     first["우편번호"],
            "주소":         addr,
            "상품명":       product_summary,
            "수량":         str(total_qty),
            "배송메시지":   first["배송메시지"],
        })

    return pd.DataFrame(rows), original_count


def match_and_fill_waybill(smart_file_obj, cj_df: pd.DataFrame):
    """
    [V3 템플릿 유지형 + 합배송 대응 송장 매칭]

    변경사항 (V3):
    - 합배송 그룹 인식: Tab 1과 동일한 정제 기준(수취인명+연락처+주소)으로 그룹화.
      그룹의 대표 주문번호(= Tab 1에서 CJ에 넘긴 고객주문번호)로 CJ 조회.
      조회된 송장번호를 그룹 내 모든 상품주문번호 행에 동일하게 기입.

    반환: (엑셀 바이트, 매칭 성공 건수, 미발급 건수, 미발급 주문번호 목록)
    """
    # ── CJ 룩업: 고객주문번호 → 운송장번호 ──
    cj_lookup: dict = {}
    for _, row in cj_df.iterrows():
        key = str(row.get("고객주문번호", "")).strip()
        val = str(row.get("운송장번호", "")).strip()
        if key and key not in cj_lookup:
            cj_lookup[key] = val

    # ── 스마트스토어 읽기 + 합배송 그룹 구성 (Tab 1과 완전 동일한 정제 기준) ──
    df_smart = read_naver_excel(smart_file_obj)

    clean_keys = pd.DataFrame({
        "order_no": df_smart.iloc[:, NAVER["상품주문번호"]].str.strip(),
        "name":     df_smart.iloc[:, NAVER["수취인명"]].apply(
                        lambda x: clean_text(str(x))),
        "phone":    df_smart.iloc[:, NAVER["수취인연락처1"]].apply(
                        lambda x: clean_phone(str(x))),
        "addr":     df_smart.iloc[:, NAVER["합배송지"]].apply(
                        lambda x: truncate_address(clean_text(str(x)))),
    })
    clean_keys = clean_keys[clean_keys["order_no"] != ""].reset_index(drop=True)

    # 각 주문번호 → 그룹 대표 주문번호 매핑
    # (같은 수취인·연락처·주소 그룹의 첫 번째 주문번호 = Tab 1의 고객주문번호)
    rep_of: dict = {}
    for _, group in clean_keys.groupby(["name", "phone", "addr"], sort=False):
        orders = group["order_no"].tolist()
        rep = orders[0]
        for o in orders:
            rep_of[o] = rep

    # 최종 맵: 주문번호 → 송장번호
    # 대표 번호로 CJ 조회, 없으면 직접 조회도 시도 (단건 주문 대응)
    order_to_waybill: dict = {}
    for order_no, rep in rep_of.items():
        waybill = cj_lookup.get(rep, "") or cj_lookup.get(order_no, "")
        if waybill:
            order_to_waybill[order_no] = waybill

    # ── 헤더 위치 기반 데이터 시작 행 계산 (동적) ──
    header_idx = find_header_row(smart_file_obj)
    data_start_row = header_idx + 2  # 0-indexed → 1-indexed(+1) → 다음 행(+1)

    # ── openpyxl 로 원본 파일 로드 (템플릿 유지) ──
    smart_file_obj.seek(0)
    wb = load_workbook(smart_file_obj)
    ws = wb.active

    matched = 0
    unmatched = 0
    unmatched_list: list = []

    for row_cells in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
        order_no = str(row_cells[NAVER["상품주문번호"]].value or "").strip()
        if not order_no:
            continue

        waybill = order_to_waybill.get(order_no, "")
        if waybill:
            row_cells[NAVER["택배사"]].value   = "CJ대한통운"
            row_cells[NAVER["송장번호"]].value = waybill
            matched += 1
        else:
            row_cells[NAVER["택배사"]].value   = "미발급"
            row_cells[NAVER["송장번호"]].value = "미발급"
            unmatched += 1
            unmatched_list.append(order_no)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), matched, unmatched, unmatched_list


def unlock_excel(file_obj, password: str = "") -> io.BytesIO:
    """
    엑셀 암호를 해제하여 BytesIO로 반환합니다.
    비밀번호가 없으면 그대로 BytesIO로 변환합니다.
    """
    file_obj.seek(0)
    raw = file_obj.read()
    if not password.strip():
        return io.BytesIO(raw)
    encrypted_buf = io.BytesIO(raw)
    office_file = msoffcrypto.OfficeFile(encrypted_buf)
    office_file.load_key(password=password.strip())
    decrypted_buf = io.BytesIO()
    office_file.decrypt(decrypted_buf)
    decrypted_buf.seek(0)
    return decrypted_buf


def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    """데이터프레임 → 엑셀 바이트 스트림 변환 (다운로드 버튼용)"""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    buf.seek(0)
    return buf.getvalue()


# ===========================================================
# Streamlit 페이지 설정
# ===========================================================
st.set_page_config(
    page_title="Summit Logic",
    page_icon="📦",
    layout="centered",
)

# ── 전역 CSS ──
st.markdown(
    """
    <style>
        .main { background-color: #ffffff; }
        body  { font-family: 'Google Sans', 'Noto Sans KR', sans-serif; }

        .header-area { text-align: center; padding: 48px 0 12px 0; }
        .header-area h1 {
            font-size: 2rem; font-weight: 700;
            color: #1a73e8; margin-bottom: 4px;
        }
        .header-area p {
            font-size: 0.95rem; color: #5f6368; line-height: 1.6;
        }

        .divider { border: none; border-top: 1px solid #e8eaed; margin: 20px 0; }

        .upload-card {
            background: #f8f9fa; border: 1px solid #e8eaed;
            border-radius: 12px; padding: 20px 24px; margin-bottom: 12px;
        }
        .upload-card h3 {
            font-size: 1rem; font-weight: 600;
            color: #202124; margin-bottom: 6px;
        }
        .upload-card p {
            font-size: 0.82rem; color: #70757a; margin-bottom: 10px;
        }

        .result-grid { display: flex; gap: 16px; margin: 20px 0; flex-wrap: wrap; }
        .stat-card {
            flex: 1; min-width: 100px;
            background: #ffffff;
            border: 1px solid #e8eaed; border-radius: 12px;
            padding: 20px 16px; text-align: center;
            box-shadow: 0 1px 3px rgba(0,0,0,0.06);
        }
        .stat-card .stat-number { font-size: 2rem; font-weight: 700; margin-bottom: 4px; }
        .stat-card .stat-label  { font-size: 0.8rem; color: #70757a; }
        .stat-total   .stat-number { color: #1a73e8; }
        .stat-matched .stat-number { color: #34a853; }
        .stat-miss    .stat-number { color: #ea4335; }
        .stat-bundle  .stat-number { color: #f9ab00; }

        .miss-box {
            background: #fff8f7; border: 1px solid #fad2cf;
            border-radius: 8px; padding: 14px 18px;
            font-size: 0.85rem; color: #c5221f;
        }

        .info-banner {
            background: #e8f0fe; border-radius: 8px;
            padding: 14px 18px; color: #1a56a4;
            font-size: 0.88rem; text-align: center; margin-top: 8px;
        }

        /* [V3] 합배송 뱃지 */
        .bundle-info {
            background: #e6f4ea; border: 1px solid #ceead6;
            border-radius: 8px; padding: 12px 16px;
            font-size: 0.85rem; color: #137333; margin: 8px 0;
        }

        /* [V3] Access Key 잠금 화면 */
        .lock-overlay {
            background: #f8f9fa; border: 1px dashed #dadce0;
            border-radius: 16px; padding: 52px 24px;
            text-align: center; color: #5f6368; margin-top: 24px;
        }
        .lock-overlay .lock-icon { font-size: 3rem; margin-bottom: 12px; }
        .lock-overlay h2 { color: #202124; font-size: 1.3rem; margin-bottom: 8px; }
        .lock-overlay p  { font-size: 0.92rem; line-height: 1.7; }

        div[data-testid="stDownloadButton"] button {
            background-color: #1a73e8; color: white;
            border: none; border-radius: 24px;
            padding: 10px 32px; font-size: 0.95rem;
            font-weight: 600; width: 100%; cursor: pointer;
            transition: background 0.2s;
        }
        div[data-testid="stDownloadButton"] button:hover {
            background-color: #1558b0;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

# ===========================================================
# [V3 신규] 사이드바: Access Key
# ===========================================================
with st.sidebar:
    st.markdown("### 🔐 Access Control")
    st.markdown("---")
    access_input = st.text_input(
        "Access Key",
        type="password",
        placeholder="접속 키를 입력하세요",
        key="access_key",
    )
    if access_input == ACCESS_KEY:
        st.success("✅ 인증 완료")
    elif access_input:
        st.error("❌ 잘못된 접속 키")
    else:
        st.info("키를 입력하면 기능이 활성화됩니다")
    st.markdown("---")
    st.caption("Summit Logic V3")

# ── 앱 헤더 ──
st.markdown(
    """
    <div class="header-area">
        <h1>📦 Summit Logic</h1>
        <p>스마트스토어 주문서와 대한통운 LOIS 파일을 업로드하면<br>
        자동으로 접수 파일 생성 및 송장번호 매칭을 처리해 드립니다.</p>
    </div>
    <hr class="divider">
    """,
    unsafe_allow_html=True,
)

# ===========================================================
# [V3] Access Key 게이트 — 인증 실패 시 이하 모든 기능 차단
# ===========================================================
if access_input != ACCESS_KEY:
    st.markdown(
        """
        <div class="lock-overlay">
            <div class="lock-icon">🔐</div>
            <h2>접근 제한</h2>
            <p>
                좌측 사이드바에 <b>Access Key</b>를 입력해야<br>
                파일 업로드 및 변환 기능을 사용할 수 있습니다.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.stop()


# ===========================================================
# 탭 레이아웃 (인증 통과 후에만 표시)
# ===========================================================
tab1, tab2 = st.tabs(["  📋 1. 접수 파일 생성  ", "  🔗 2. 송장 번호 매칭  "])


# ===========================================================
# 탭 1: 접수 파일 생성
# ===========================================================
with tab1:

    st.markdown("#### 대한통운 LOIS 접수 파일 생성")
    st.markdown(
        """
        <div class="info-banner">
            네이버 스마트스토어 주문서를 올리면 CJ 대한통운 LOIS 업로드 전용 양식으로 변환합니다.<br>
            <small>스마트스토어 &gt; 발주(주문)확인/발송관리 &gt; 엑셀 다운로드 파일을 사용하세요.</small>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown(
        """
        <div class="upload-card">
            <h3>① 스마트스토어 주문서 업로드</h3>
            <p>네이버 스마트스토어에서 다운로드한 주문 엑셀 파일을 올려주세요.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    uploaded_t1 = st.file_uploader(
        "스마트스토어 주문서 (xlsx)",
        type=["xlsx"],
        key="tab1_upload",
        label_visibility="collapsed",
    )
    pw_t1 = st.text_input(
        "Excel Password (Optional)",
        type="password",
        key="tab1_pw",
        placeholder="엑셀 파일에 비밀번호가 있는 경우에만 입력하세요",
    )

    if uploaded_t1:
        try:
            unlocked_t1 = unlock_excel(uploaded_t1, pw_t1)
            df_smart = read_naver_excel(unlocked_t1)

            # [V3] build_cj_upload_df 는 (df, original_count) 튜플 반환
            df_cj_upload, original_count = build_cj_upload_df(df_smart)
            total = len(df_cj_upload)
            bundled = original_count - total  # 합배송으로 절약된 건수

            # ── 결과 통계 카드 ──
            bundle_html = (
                f'<div class="stat-card stat-bundle">'
                f'<div class="stat-number">{bundled}</div>'
                f'<div class="stat-label">🔗 합배송 절약 건</div>'
                f'</div>'
            ) if bundled > 0 else ""

            st.markdown(
                f"""
                <div class="result-grid">
                    <div class="stat-card stat-total">
                        <div class="stat-number">{original_count}</div>
                        <div class="stat-label">원본 주문 건수</div>
                    </div>
                    <div class="stat-card stat-matched">
                        <div class="stat-number">{total}</div>
                        <div class="stat-label">✅ 발송 건수</div>
                    </div>
                    {bundle_html}
                </div>
                """,
                unsafe_allow_html=True,
            )

            # 합배송 안내 메시지
            if bundled > 0:
                st.markdown(
                    f"""
                    <div class="bundle-info">
                        🔗 <b>합배송 {bundled}건 자동 감지</b> — 수취인·연락처·주소가 동일한 주문을
                        1건으로 묶었습니다. 상품명은 <code>상품A 외 N건</code> 형태로 요약되었습니다.
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

            # 컬럼 매핑 안내
            with st.expander("컬럼 매핑 확인"):
                st.table(pd.DataFrame({
                    "스마트스토어 컬럼": [
                        "A열 상품주문번호", "N열 수취인명", "AW열 수취인연락처1",
                        "BC열 우편번호", "AY열 합배송지", "U열 상품명",
                        "AA열 수량", "BD열 배송메세지",
                    ],
                    "→ CJ LOIS 컬럼": [
                        "고객주문번호", "수취인명 (이모지 제거)",
                        "연락처 (숫자만)", "우편번호",
                        "주소 (이모지 제거, 100자 제한)", "상품명 (합배송 요약)",
                        "수량 (합산)", "배송메시지 (이모지 제거)",
                    ],
                }))

            with st.expander("📋 변환 결과 미리보기", expanded=True):
                st.dataframe(df_cj_upload, use_container_width=True)

            st.markdown("<br>", unsafe_allow_html=True)
            st.download_button(
                label="⬇️  CJ LOIS 접수 파일 다운로드",
                data=df_to_excel_bytes(df_cj_upload, "LOIS_접수"),
                file_name="CJ_LOIS_접수.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        except IndexError:
            st.error("컬럼 구조가 예상과 다릅니다. 네이버 스마트스토어 원본 엑셀 파일인지 확인해 주세요.")
        except Exception as e:
            st.error(f"처리 중 오류가 발생했습니다: {e}")
            with st.expander("오류 상세"):
                st.exception(e)

    else:
        st.markdown(
            """
            <div class="info-banner" style="margin-top:16px;">
                📂 스마트스토어 주문서 파일을 업로드해 주세요.
            </div>
            """,
            unsafe_allow_html=True,
        )


# ===========================================================
# 탭 2: 송장 번호 매칭
# ===========================================================
with tab2:

    st.markdown("#### 대한통운 → 스마트스토어 송장번호 자동 매칭")
    st.markdown(
        """
        <div class="info-banner">
            두 파일을 올리면 <b>상품주문번호 ↔ 고객주문번호</b> 기준으로 자동 매칭하여<br>
            H열(택배사)과 I열(송장번호)을 채운 파일을 반환합니다.<br>
            <small>합배송 묶음 주문은 동일한 송장번호가 모든 관련 행에 자동 입력됩니다.</small>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("<br>", unsafe_allow_html=True)

    col_l, col_r = st.columns(2)

    with col_l:
        st.markdown(
            """
            <div class="upload-card">
                <h3>① 스마트스토어 원본 파일</h3>
                <p>네이버 스마트스토어 주문 엑셀 원본 파일을 올려주세요.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        uploaded_smart_t2 = st.file_uploader(
            "스마트스토어 원본 (xlsx)",
            type=["xlsx"],
            key="tab2_smart",
            label_visibility="collapsed",
        )
        pw_t2 = st.text_input(
            "Excel Password (Optional)",
            type="password",
            key="tab2_pw",
            placeholder="비밀번호가 있는 경우에만 입력",
        )

    with col_r:
        st.markdown(
            """
            <div class="upload-card">
                <h3>② 대한통운 LOIS 결과 파일</h3>
                <p>LOIS 시스템에서 운송장 발급 후 다운로드한 결과 파일을 올려주세요.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        uploaded_cj_t2 = st.file_uploader(
            "대한통운 LOIS 결과 (xlsx)",
            type=["xlsx"],
            key="tab2_cj",
            label_visibility="collapsed",
        )

    st.markdown("<br>", unsafe_allow_html=True)
    run_btn = st.button("🔍 송장번호 자동 매칭 실행", use_container_width=True, key="run_btn")
    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    if run_btn:
        if not uploaded_smart_t2 or not uploaded_cj_t2:
            missing = []
            if not uploaded_smart_t2: missing.append("스마트스토어 원본 파일 ①")
            if not uploaded_cj_t2:    missing.append("대한통운 LOIS 결과 파일 ②")
            st.markdown(
                f'<div class="info-banner">📂 <b>{", ".join(missing)}</b>를 먼저 업로드해주세요.</div>',
                unsafe_allow_html=True,
            )
        else:
            try:
                with st.spinner("매칭 처리 중... 잠시만 기다려주세요."):
                    unlocked_smart_t2 = unlock_excel(uploaded_smart_t2, pw_t2)

                    df_cj = pd.read_excel(uploaded_cj_t2, dtype=str).fillna("")
                    required_cj = ["고객주문번호", "운송장번호"]
                    missing_cols = [c for c in required_cj if c not in df_cj.columns]
                    if missing_cols:
                        raise ValueError(
                            f"대한통운 파일에 필수 컬럼이 없습니다: {missing_cols}\n"
                            f"실제 컬럼: {list(df_cj.columns)}"
                        )

                    result_bytes, matched, unmatched, unmatched_list = match_and_fill_waybill(
                        smart_file_obj=unlocked_smart_t2,
                        cj_df=df_cj,
                    )

                total = matched + unmatched

                st.markdown("### 📊 매칭 결과 요약")
                st.markdown(
                    f"""
                    <div class="result-grid">
                        <div class="stat-card stat-total">
                            <div class="stat-number">{total}</div>
                            <div class="stat-label">전체 주문 건수</div>
                        </div>
                        <div class="stat-card stat-matched">
                            <div class="stat-number">{matched}</div>
                            <div class="stat-label">✅ 매칭 성공</div>
                        </div>
                        <div class="stat-card stat-miss">
                            <div class="stat-number">{unmatched}</div>
                            <div class="stat-label">❌ 미발급</div>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                if unmatched_list:
                    miss_html = "<br>".join(f"• {o}" for o in unmatched_list)
                    st.markdown(
                        f"""
                        <div class="miss-box">
                            <b>⚠ 미발급 주문번호 목록</b><br><br>{miss_html}
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )
                    st.markdown("<br>", unsafe_allow_html=True)
                else:
                    st.success("모든 주문의 송장번호가 성공적으로 매칭되었습니다!")

                # ── 결과 미리보기 ──
                header_row_prev = find_header_row(unlocked_smart_t2)
                unlocked_smart_t2.seek(0)
                df_preview = pd.read_excel(
                    unlocked_smart_t2, header=header_row_prev, dtype=str
                ).fillna("")

                cj_lookup_prev = dict(
                    zip(df_cj["고객주문번호"].str.strip(), df_cj["운송장번호"].str.strip())
                )

                # Tab 2 미리보기에서도 합배송 그룹 대응
                clean_keys_prev = pd.DataFrame({
                    "order_no": df_preview.iloc[:, NAVER["상품주문번호"]].str.strip(),
                    "name":     df_preview.iloc[:, NAVER["수취인명"]].apply(lambda x: clean_text(str(x))),
                    "phone":    df_preview.iloc[:, NAVER["수취인연락처1"]].apply(lambda x: clean_phone(str(x))),
                    "addr":     df_preview.iloc[:, NAVER["합배송지"]].apply(lambda x: truncate_address(clean_text(str(x)))),
                })
                rep_of_prev: dict = {}
                for _, grp in clean_keys_prev.groupby(["name", "phone", "addr"], sort=False):
                    ords = grp["order_no"].tolist()
                    r = ords[0]
                    for o in ords:
                        rep_of_prev[o] = r

                preview = df_preview.iloc[:, [
                    NAVER["상품주문번호"], NAVER["수취인명"],
                    NAVER["상품명"], NAVER["택배사"], NAVER["송장번호"],
                ]].copy()
                preview.columns = ["상품주문번호", "수취인명", "상품명", "택배사", "송장번호"]
                preview = preview[preview["상품주문번호"].str.strip() != ""].copy()

                for i, row in preview.iterrows():
                    key = str(row["상품주문번호"]).strip()
                    rep = rep_of_prev.get(key, key)
                    wb_no = cj_lookup_prev.get(rep, "") or cj_lookup_prev.get(key, "")
                    preview.at[i, "택배사"]  = "CJ대한통운" if wb_no else "미발급"
                    preview.at[i, "송장번호"] = wb_no if wb_no else "미발급"

                with st.expander("📋 결과 미리보기", expanded=False):
                    st.dataframe(preview, use_container_width=True)

                st.markdown("<br>", unsafe_allow_html=True)
                st.download_button(
                    label="⬇️  송장번호 기입 완료된 스마트스토어 파일 다운로드",
                    data=result_bytes,
                    file_name="스마트스토어_송장입력완료.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.caption(
                    "다운로드한 파일을 스마트스토어 "
                    "'발주(주문)확인/발송관리 > 일괄발송 처리' 메뉴에서 업로드하세요."
                )

            except ValueError as ve:
                st.error(f"파일 형식 오류\n\n{ve}")
            except Exception as e:
                st.error(f"처리 중 오류가 발생했습니다: {e}")
                with st.expander("오류 상세"):
                    st.exception(e)

    else:
        if uploaded_smart_t2 and uploaded_cj_t2:
            st.markdown(
                '<div class="info-banner">✅ 두 파일이 모두 업로드되었습니다. <b>매칭 실행 버튼</b>을 눌러주세요.</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div class="info-banner">📂 위에서 두 파일을 모두 업로드한 뒤 매칭 실행 버튼을 눌러주세요.</div>',
                unsafe_allow_html=True,
            )


# ===========================================================
# 하단 푸터
# ===========================================================
st.markdown("<br><br>", unsafe_allow_html=True)
st.markdown(
    """
    <div style="text-align:center; color:#bdc1c6; font-size:0.78rem;">
        Summit Logic V3 &nbsp;|&nbsp; 스마트스토어 × 대한통운 LOIS 자동화
    </div>
    """,
    unsafe_allow_html=True,
)
