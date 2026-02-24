# -*- coding: utf-8 -*-
"""
logistics_engine.py
───────────────────
스마트스토어 × 대한통운 LOIS 핵심 물류 처리 엔진.

공개 API:
  NAVER                 : 스마트스토어 엑셀 컬럼 인덱스 상수 딕셔너리
  find_header_row()     : 엑셀에서 '상품주문번호' 헤더 행 번호 탐색
  read_naver_excel()    : 네이버 엑셀 안전 읽기 (불량 행 자동 제거)
  build_cj_upload_df()  : 스마트스토어 → CJ LOIS 접수 양식 변환 (합배송 + 정제)
  match_and_fill_waybill(): 송장번호 매칭 후 원본 템플릿에 기입 (합배송 대응)
  df_to_excel_bytes()   : DataFrame → 다운로드용 엑셀 바이트 변환
"""

import io
import pandas as pd
from openpyxl import load_workbook

from data_cleaner import clean_text, clean_phone, truncate_address


# ===========================================================
# 상수: 네이버 스마트스토어 엑셀 컬럼 인덱스 (0-based, 헤더 행 기준)
# ===========================================================
NAVER: dict[str, int] = {
    "상품주문번호":  0,   # A열 — 매칭 기준 키
    "택배사":       7,   # H열 — 채워야 할 칸
    "송장번호":     8,   # I열 — 채워야 할 칸
    "수취인명":     13,  # N열
    "상품명":       20,  # U열
    "수량":         26,  # AA열
    "수취인연락처1": 48,  # AW열
    "합배송지":     50,  # AY열 (기본주소 + 상세주소 합본)
    "우편번호":     54,  # BC열
    "배송메세지":   55,  # BD열
}


# ===========================================================
# [V3.1] 지능형 컬럼 탐색
# ===========================================================

# ── CJ LOIS 파일 컬럼 탐색 키워드 (우선순위 순, 한·영 혼용 지원) ──
_CJ_ORDER_KEYWORDS: list[str] = [
    "고객주문번호", "주문번호", "고객주문", "주문",
    "order number", "order no", "orderno", "order",   # 영문 양식 대응
]
_CJ_WAYBILL_KEYWORDS: list[str] = [
    "운송장번호", "송장번호", "운송장", "invoice", "송장", "운송",
    "waybill", "tracking number", "tracking no", "tracking",  # 영문 양식 대응
]

# ── 스마트스토어 파일에서 '상품주문번호' 컬럼 탐색 키워드 ──
_SMART_ORDER_KEYWORDS: list[str] = [
    "상품주문번호", "주문번호", "상품주문",
]


def find_column(df: pd.DataFrame, keywords: list, field_name: str) -> str:
    """
    DataFrame 컬럼 목록에서 키워드와 일치하거나 포함하는 컬럼명을 탐색합니다.

    탐색 순서:
    1. keywords 순서대로 정확히 일치하는 컬럼 탐색 (strip 후 비교)
    2. 정확 일치가 없으면 keywords 순서대로 포함(contains) 탐색 (대소문자 무시)

    Args:
        df         : 검색 대상 DataFrame
        keywords   : 우선순위 순 탐색 키워드 리스트
        field_name : 오류 메시지에 표시할 필드 이름

    Returns:
        발견된 실제 컬럼명 (str)

    Raises:
        ValueError: 매칭되는 컬럼이 없을 경우 구체적인 오류 메시지 포함.
    """
    cols = [str(c) for c in df.columns]

    # 1단계: 정확히 일치
    for kw in keywords:
        for col in cols:
            if col.strip() == kw:
                return col

    # 2단계: 포함 일치 (대소문자 무시)
    for kw in keywords:
        for col in cols:
            if kw.lower() in col.strip().lower():
                return col

    raise ValueError(
        f"파일 양식이 잘못되었습니다. '{field_name}' 컬럼을 찾을 수 없습니다.\n"
        f"탐색 키워드: {keywords}\n"
        f"실제 컬럼 목록: {cols}"
    )


def map_cj_columns(df: pd.DataFrame) -> dict[str, str]:
    """
    CJ LOIS 결과 파일에서 주문번호·운송장번호 컬럼을 지능적으로 탐색합니다.

    '운송장번호'뿐만 아니라 '송장번호', '운송장', 'invoice' 등
    유사 명칭도 자동으로 인식합니다.

    Returns:
        dict: {"order": 실제_주문번호_컬럼명, "waybill": 실제_운송장번호_컬럼명}

    Raises:
        ValueError: 필수 컬럼을 찾지 못한 경우 구체적인 오류 메시지 포함.
    """
    order_col   = find_column(df, _CJ_ORDER_KEYWORDS,   "고객주문번호 (주문 키)")
    waybill_col = find_column(df, _CJ_WAYBILL_KEYWORDS, "운송장번호 (송장 번호)")
    return {"order": order_col, "waybill": waybill_col}


def diagnose_smart_file(df_smart: pd.DataFrame, header_row: int) -> dict:
    """
    스마트스토어 DataFrame의 컬럼 인식 결과를 진단 정보로 반환합니다.

    Returns:
        dict: {
            "header_row"   : 헤더 행 번호 (0-indexed),
            "total_cols"   : 전체 컬럼 수,
            "total_rows"   : 데이터 행 수,
            "key_cols"     : {논리명: (인덱스, 실제컬럼명, 인식여부)} 딕셔너리,
        }
    """
    key_map = {
        "상품주문번호":  NAVER["상품주문번호"],
        "수취인명":     NAVER["수취인명"],
        "수취인연락처1": NAVER["수취인연락처1"],
        "합배송지":     NAVER["합배송지"],
        "상품명":       NAVER["상품명"],
        "수량":         NAVER["수량"],
        "택배사":       NAVER["택배사"],
        "송장번호":     NAVER["송장번호"],
    }
    key_cols = {}
    for logical, idx in key_map.items():
        if idx < len(df_smart.columns):
            actual = str(df_smart.columns[idx])
            ok = logical in actual or actual in logical
            key_cols[logical] = (idx, actual, ok)
        else:
            key_cols[logical] = (idx, "컬럼 없음", False)

    return {
        "header_row": header_row,
        "total_cols": len(df_smart.columns),
        "total_rows": len(df_smart),
        "key_cols":   key_cols,
    }


# ===========================================================
# 헤더 탐색 & 엑셀 읽기
# ===========================================================

def find_header_row(file_obj) -> int:
    """
    '상품주문번호' 텍스트와 정확히 일치하는 셀이 있는 행 번호(0-based)를 반환합니다.

    네이버 엑셀 상단의 안내 문구에도 '상품주문번호' 단어가 포함되므로,
    contains() 대신 == 비교를 사용해 안내 문구 행을 완전히 배제합니다.

    Raises:
        ValueError: '상품주문번호' 컬럼을 찾지 못한 경우.
    """
    file_obj.seek(0)
    df_raw = pd.read_excel(file_obj, header=None, dtype=str)
    for idx, row in df_raw.iterrows():
        if (row.astype(str).str.strip() == "상품주문번호").any():
            return int(idx)
    raise ValueError(
        "'상품주문번호' 컬럼을 찾을 수 없습니다.\n"
        "네이버 스마트스토어에서 다운로드한 원본 엑셀 파일인지 확인해 주세요."
    )


def read_naver_excel(file_obj) -> pd.DataFrame:
    """
    네이버 스마트스토어 주문 엑셀을 안전하게 읽습니다.

    처리 순서:
    1. find_header_row() 로 헤더 행 동적 탐색
    2. 해당 행을 컬럼 헤더로 설정해 DataFrame 생성
    3. 불량 행 제거
       - 상품주문번호 열이 빈 행 (빈 줄·합계 행 등)
       - 상품주문번호 열이 '상품주문번호' 텍스트 그 자체인 행 (중복 헤더 잔재)
    4. dtype=str → 주문번호·전화번호 앞자리 0 보존
    """
    header_row = find_header_row(file_obj)
    file_obj.seek(0)
    df = pd.read_excel(file_obj, header=header_row, dtype=str)
    df = df.fillna("")

    order_col = df.columns[0]
    df = df[
        (df[order_col].str.strip() != "") &
        (df[order_col].str.strip() != "상품주문번호")
    ].reset_index(drop=True)
    return df


# ===========================================================
# 접수 파일 생성 (Tab 1)
# ===========================================================

def build_cj_upload_df(df_smart: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """
    스마트스토어 DataFrame → CJ 대한통운 LOIS 접수 양식 변환.

    V3 처리:
    1. 데이터 정제 (data_cleaner 모듈 사용)
       - 이름·주소·배송메시지 : 이모지·제어문자 제거 (clean_text)
       - 전화번호             : 숫자만 추출 (clean_phone)
       - 주소                 : 100자 초과 시 절삭 (truncate_address)
    2. 합배송 그룹핑 (수취인명 + 연락처 + 주소 기준)
       - 동일 수취인·연락처·주소 → 1건으로 묶음
       - 상품명: "상품A 외 N건" 형태 요약
       - 수량: 그룹 내 합산
       - 고객주문번호: 그룹 첫 번째 상품주문번호 (Tab 2 매칭 키)

    Returns:
        tuple: (변환된 DataFrame, 원본 주문 건수)
    """
    # ── [V3.1] 데이터 유효성 검사 ──
    # 스마트스토어 파일의 컬럼 수가 필요한 최소 인덱스보다 적으면 즉시 오류 안내
    required_max_idx = max(NAVER.values())
    if df_smart.shape[1] <= required_max_idx:
        raise ValueError(
            f"파일 양식이 잘못되었습니다. '상품주문번호' 컬럼을 찾을 수 없습니다.\n"
            f"예상 컬럼 수: {required_max_idx + 1}개 이상 / "
            f"실제 컬럼 수: {df_smart.shape[1]}개\n"
            "네이버 스마트스토어에서 다운로드한 원본 엑셀 파일인지 확인해 주세요."
        )

    # '상품주문번호' 컬럼이 이름으로도 존재하는지 추가 확인
    try:
        find_column(df_smart, _SMART_ORDER_KEYWORDS, "상품주문번호")
    except ValueError:
        # 이름으로 못 찾아도 positional 접근이 있으므로 경고만 (raise 안 함)
        pass

    # ── 1) 컬럼 추출 + 정제 ──
    df = pd.DataFrame({
        "고객주문번호": df_smart.iloc[:, NAVER["상품주문번호"]].str.strip(),
        "수취인명":     df_smart.iloc[:, NAVER["수취인명"]].apply(
                            lambda x: clean_text(str(x))),
        "연락처":       df_smart.iloc[:, NAVER["수취인연락처1"]].apply(
                            lambda x: clean_phone(str(x))),
        "우편번호":     df_smart.iloc[:, NAVER["우편번호"]].str.strip(),
        "주소":         df_smart.iloc[:, NAVER["합배송지"]].apply(
                            lambda x: truncate_address(clean_text(str(x)))),
        "상품명":       df_smart.iloc[:, NAVER["상품명"]].str.strip(),
        "수량":         df_smart.iloc[:, NAVER["수량"]].str.strip(),
        "배송메시지":   df_smart.iloc[:, NAVER["배송메세지"]].apply(
                            lambda x: clean_text(str(x))),
    })
    df = df[df["고객주문번호"] != ""].reset_index(drop=True)
    original_count = len(df)

    # ── 2) 합배송 그룹핑 ──
    rows = []
    for (name, phone, addr), group in df.groupby(
        ["수취인명", "연락처", "주소"], sort=False
    ):
        first = group.iloc[0]
        products = group["상품명"].tolist()
        product_summary = (
            products[0]
            if len(products) == 1
            else f"{products[0]} 외 {len(products) - 1}건"
        )
        try:
            qty_list = [int(q) for q in group["수량"] if str(q).strip().isdigit()]
            total_qty = sum(qty_list) if qty_list else first["수량"]
        except Exception:
            total_qty = first["수량"]

        rows.append({
            "고객주문번호": first["고객주문번호"],
            "수취인명":     name,
            "연락처":       phone,
            "우편번호":     first["우편번호"],
            "주소":         addr,
            "상품명":       product_summary,
            "수량":         str(total_qty),
            "배송메시지":   first["배송메시지"],
        })

    return pd.DataFrame(rows), original_count


# ===========================================================
# 송장 매칭 (Tab 2)
# ===========================================================

def match_and_fill_waybill(
    smart_file_obj,
    cj_df: pd.DataFrame,
) -> tuple[bytes, int, int, list[str]]:
    """
    [템플릿 유지형 + 합배송 대응] 스마트스토어 원본에 송장번호를 기입합니다.

    처리 방식:
    1. CJ 파일에서 {고객주문번호: 운송장번호} 룩업 사전 생성
    2. 스마트스토어 파일을 read_naver_excel() 로 읽어 합배송 그룹 구성
       (Tab 1과 완전 동일한 정제 기준 → 그룹이 일치함을 보장)
    3. 그룹 대표 주문번호 → CJ 조회 → 그룹 내 모든 행에 동일 송장번호 기입
    4. openpyxl 로 원본 파일 로드 → H열(택배사)·I열(송장번호) 셀만 수정
       (1·2행 안내 문구, 서식, 수식 등 모든 원본 내용 보존)

    Returns:
        tuple: (수정된 엑셀 바이트, 매칭 성공 건수, 미발급 건수, 미발급 주문번호 목록)
    """
    # ── [V3.1] 지능형 컬럼 탐색으로 CJ 룩업 생성 ──
    # '운송장번호', '송장번호', 'invoice' 등 유사 명칭 컬럼 자동 인식
    cj_col_map  = map_cj_columns(cj_df)          # 데이터 유효성 검사 + 컬럼 탐색
    order_col   = cj_col_map["order"]
    waybill_col = cj_col_map["waybill"]

    cj_lookup: dict[str, str] = {}
    for _, row in cj_df.iterrows():
        key = str(row.get(order_col, "")).strip()
        val = str(row.get(waybill_col, "")).strip()
        if key and key not in cj_lookup:
            cj_lookup[key] = val

    # ── 합배송 그룹 구성 (Tab 1과 동일 기준) ──
    df_smart = read_naver_excel(smart_file_obj)

    clean_keys = pd.DataFrame({
        "order_no": df_smart.iloc[:, NAVER["상품주문번호"]].str.strip(),
        "name":     df_smart.iloc[:, NAVER["수취인명"]].apply(
                        lambda x: clean_text(str(x))),
        "phone":    df_smart.iloc[:, NAVER["수취인연락처1"]].apply(
                        lambda x: clean_phone(str(x))),
        "addr":     df_smart.iloc[:, NAVER["합배송지"]].apply(
                        lambda x: truncate_address(clean_text(str(x)))),
    })
    clean_keys = clean_keys[clean_keys["order_no"] != ""].reset_index(drop=True)

    # 주문번호 → 그룹 대표 주문번호 매핑
    rep_of: dict[str, str] = {}
    for _, group in clean_keys.groupby(["name", "phone", "addr"], sort=False):
        orders = group["order_no"].tolist()
        rep = orders[0]
        for o in orders:
            rep_of[o] = rep

    # 최종 맵: 주문번호 → 송장번호 (대표 번호 우선, 없으면 직접 조회)
    order_to_waybill: dict[str, str] = {}
    for order_no, rep in rep_of.items():
        waybill = cj_lookup.get(rep, "") or cj_lookup.get(order_no, "")
        if waybill:
            order_to_waybill[order_no] = waybill

    # ── 헤더 위치 기반 데이터 시작 행 계산 ──
    header_idx = find_header_row(smart_file_obj)
    data_start_row = header_idx + 2  # 0-indexed → 1-indexed(+1) → 다음 행(+1)

    # ── openpyxl 로 원본 파일 로드 (템플릿 유지) ──
    smart_file_obj.seek(0)
    wb = load_workbook(smart_file_obj)
    ws = wb.active

    matched = 0
    unmatched = 0
    unmatched_list: list[str] = []

    for row_cells in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
        order_no = str(row_cells[NAVER["상품주문번호"]].value or "").strip()
        if not order_no:
            continue
        waybill = order_to_waybill.get(order_no, "")
        if waybill:
            row_cells[NAVER["택배사"]].value   = "CJ대한통운"
            row_cells[NAVER["송장번호"]].value = waybill
            matched += 1
        else:
            row_cells[NAVER["택배사"]].value   = "미발급"
            row_cells[NAVER["송장번호"]].value = "미발급"
            unmatched += 1
            unmatched_list.append(order_no)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), matched, unmatched, unmatched_list


# ===========================================================
# 공통 유틸
# ===========================================================

def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    """
    DataFrame을 다운로드 버튼용 엑셀 바이트 스트림으로 변환합니다.

    Args:
        df         : 변환할 DataFrame
        sheet_name : 엑셀 시트 이름 (기본값: "Sheet1")

    Returns:
        bytes: openpyxl 엔진으로 생성된 .xlsx 파일 바이트
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    buf.seek(0)
    return buf.getvalue()
